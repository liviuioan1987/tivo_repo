import pandas
import csv
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import re
import pandas
from openpyxl.workbook import Workbook

def read_mso_localities(packaged_service_id,operator_file, pluto="yes"):
    global mso_locality_template
    tivo_plus_localities = []
    file = pandas.ExcelFile(operator_file)
    all_localities = file.sheet_names

    reading_page = file.parse("Channel")
    operator_channel_page_query = reading_page.query(f'{packaged_service_id} <= `Packaged Service Id`  < {packaged_service_id + 300}')
    operator_channel_page = operator_channel_page_query["Packaged Service Id"].size

    for locality in all_localities[1:]:
        reading_pages  = file.parse(locality)
        try:
            reading_pages["Packaged Service Id"] = pandas.to_numeric(reading_pages["Packaged Service Id"],downcast='integer', errors='coerce')
            operator_locality_page_query = reading_pages.query(f'{packaged_service_id} <= `Packaged Service Id`  < {packaged_service_id + 300}')
            operator_locality_page = operator_locality_page_query["Packaged Service Id"].size
            if operator_locality_page > 0:
                if operator_locality_page != operator_channel_page:
                    print(f" In locality {locality} not all linear channels were present, there are {operator_locality_page} and were expected {operator_channel_page}")
                    if "-" in locality:
                        locality_split = locality.split("-")
                        tivo_plus_localities.append(locality_split[1])
                    else:
                        tivo_plus_localities.append(locality)

                elif operator_locality_page == operator_channel_page:
                    if "-" in locality:
                        locality_split = locality.split("-")
                        tivo_plus_localities.append(locality_split[1])
                        mso_locality_template = locality
                    else:
                        tivo_plus_localities.append(locality)
                        mso_locality_template = locality
        except:
            pass
    return tivo_plus_localities

class Xumo_processor():
    def __init__(self,solution_source_file,xumo_partner_id,list_msoservice_IDs,packaged_service_id,mso_all_data_file):
        self.workbook = Workbook()
        self.solution_source_file = solution_source_file
        self.channel_start_range = 1
        self.list_msoservice_IDs = list_msoservice_IDs
        self.xumo_partner_id = xumo_partner_id
        self.mso_all_data_file = mso_all_data_file
        self.packaged_service_id = packaged_service_id

    # Regex strip pattern
    pattern = re.compile(r';"(.+)"')

    def xumo_csv_to_data(self):
        file = open(self.solution_source_file, "rt", encoding="utf-16")
        open_csv_file = csv.reader(file, delimiter='\t')

        extracted_data = [[row[1], row[8], row[10], row[2]] for row in open_csv_file if row[0].isdigit()]
        list_confluence_channels = [{ "stationId": "epgProvider:st." + entry[1],"url": entry[2]} for entry in extracted_data]

        # Adding LCN to the linear channels
        for channel in list_confluence_channels:
            channel["linear_channel_nr"] = self.channel_start_range
            self.channel_start_range += 1

        # Removing empty elements from the list
        pattern_stationid = re.compile(r'epgProvider:st.(\d+)')
        self.list_confluence_channels_stripped = [item for item in list_confluence_channels if len(pattern_stationid.findall(item["stationId"])) >= 1]

        # Removing the =URL formatting if found at channel level
        # Regex strip pattern
        pattern = re.compile(r';"(.+)"')
        for channel in self.list_confluence_channels_stripped:
            url_search = pattern.findall(channel["url"])
            if url_search:
                channel["url"] = url_search[0]

            else:
                pass
        # Only custom rule setup as Amanda keeps leaving the channel as it is
        for channel in self.list_confluence_channels_stripped:
            if channel["stationId"] == "epgProvider:st.10427289051":
                channel["stationId"] = "epgProvider:st.446201192"

        return self.list_confluence_channels_stripped

    def open_mso_file_data(self):
        packaged_service_id = self.packaged_service_id
        data = pandas.read_excel(self.mso_all_data_file)
        all_data = data.loc[:,["Packaged Service Id", "Partner Station Id", "Call Sign", "Station Name", "Packaged Service Description", "Availability Window Start","Availability Window End","IP ABR URL"]]
        y = all_data.query(f'{packaged_service_id} <= `Packaged Service Id`  < {packaged_service_id + 400}')
        self.pckd_serv_id = list(y["Packaged Service Id"])
        prtn_st_id = list(y["Partner Station Id"])
        call_sign = list(y["Call Sign"])
        channel_st_id = list(y["Station Name"])
        packaged_service_description = list(y["Packaged Service Description"])
        start_date = list(y["Availability Window Start"])
        end_date = list(y["Availability Window End"])
        url = list(y["IP ABR URL"])
        self.validation_data = [
            {"channelName": channel_st_id[i], "stationId": prtn_st_id[i], "packaged_service_id": self.pckd_serv_id[i], "packaged_service_description": packaged_service_description[i], "start_date": start_date[i], "end_date": end_date[i], "url": url[i], "call_sign":call_sign[i], "lcn":None} for i in range(len(prtn_st_id))]

        ## findinding the LCN for the channels
        data2 = pandas.read_excel(self.mso_all_data_file, sheet_name=mso_locality_template)
        all_data = data2.loc[:,["Packaged Service Id","Partner Station Id","Logical Channel Number"]]
        all_data["Packaged Service Id"] = pandas.to_numeric(all_data["Packaged Service Id"], downcast='integer', errors='coerce')
        y = all_data.query(f'{packaged_service_id} <= `Packaged Service Id`  < {packaged_service_id + 400}')
        prtn_st_id = list(y["Partner Station Id"])
        lcn = list(y["Logical Channel Number"])

        self.validation_data2 = [{"stationId": prtn_st_id[i], "lcn":int(lcn[i])} for i in range(len(prtn_st_id))]

        for station in self.validation_data:
            for lcns in self.validation_data2:
                if station["stationId"] == lcns["stationId"]:
                    station["lcn"] = lcns["lcn"]

        return self.validation_data, self.pckd_serv_id

    def compare_and_filter(self):
        for operator_channel in self.validation_data:
            for conf_channel in self.list_confluence_channels_stripped:
                if operator_channel["stationId"] == conf_channel["stationId"]:
                    operator_channel["url"] = conf_channel["url"]

        return self.validation_data

    def write_channel_page(self):
        channel_sheet = self.workbook["Sheet"]
        channel_sheet.title = "Channel"

        header_data = ["Packaged Service Id", "Partner Station Id", "Call Sign", "Station Name","Packaged Service Description", "Jump Channel Type", "Availability Window Start", \
                       "Availability Window End", "Linear Service Type", "DVB Linkage Id", "Linear Provider Partner Id","Recording Provider Partner Id", \
                       "Partner Channel Id", "Video Resolution", "Prevent EAS Interruption", "Channel Change Id", \
                       "Playback URI Channel Id", "IP ABR URL", "DRM Type", "Transport Encoding Type", "Device Type","SOCU Base URL", "Application Id", "Channel Name", \
                       "Channel Description", "Guide Cell Title", "Logo Partner Id", "DRM Content ID", "Multi View"]

        incremental_column = 1

        for value in range(0, len(header_data)):
            channel_sheet.cell(row=1, column=incremental_column).value = header_data[value]
            channel_sheet.cell(row=1, column=incremental_column).fill = PatternFill(start_color="00CCFF",
                                                                                    end_color="00CCFF",
                                                                                    fill_type="solid")
            channel_sheet.column_dimensions[get_column_letter(incremental_column)].width = 25
            incremental_column += 1

        incremental_row = 2
        for entry in self.validation_data:
            channel_sheet.cell(row=incremental_row, column=1).value = entry["packaged_service_id"]
            channel_sheet.cell(row=incremental_row, column=2).value = entry["stationId"]
            channel_sheet.cell(row=incremental_row, column=3).value = entry["call_sign"]

            channel_sheet.cell(row=incremental_row, column=4).value = entry["channelName"]
            channel_sheet.cell(row=incremental_row,column=5).value = entry["packaged_service_description"]

            channel_sheet.cell(row=incremental_row, column=7).value = entry["start_date"]
            channel_sheet.cell(row=incremental_row, column=8).value = entry["end_date"]
            channel_sheet.cell(row=incremental_row, column=9).value = "linear"
            channel_sheet.cell(row=incremental_row, column=11).value = f"tivo:pt.{self.xumo_partner_id}"

            channel_sheet.cell(row=incremental_row, column=15).value = "false"

            channel_sheet.cell(row=incremental_row, column=18).value = entry["url"]
            channel_sheet.cell(row=incremental_row, column=20).value = "hlsTransportStream"

            incremental_row += 1

    def write_msoSiteID_pages(self):
        #self.confluence_file, list_msoservice_IDs
        page_columns = ["Packaged Service Id", "Partner Station Id", "Call Sign", "Station Name","Packaged Service Description", "Partner Channel Id", \
                        "Logical Channel Number", "IP ABR URL", "Channel Change Id", "Playback URI Channel Id", "Transport Encoding Type", "Preferred Transport", "Device Type", "DRM Type", \
                        "SOCU Base URL", "Application Id", "Channel Name", "Channel Description", "Guide Cell Title","Logo Partner Id" \
                        ]
        primary_page = self.list_msoservice_IDs[0]
        self.workbook.create_sheet(primary_page)
        channel_sheet = self.workbook[primary_page]

        incremental_column = 1
        for data in range(len(page_columns)):
            channel_sheet.cell(row=1, column=incremental_column).value = page_columns[data]
            channel_sheet.cell(row=1, column=incremental_column).fill = PatternFill(start_color="00CCFF",
                                                                                    end_color="00CCFF",
                                                                                    fill_type="solid")
            channel_sheet.column_dimensions[get_column_letter(incremental_column)].width = 25
            incremental_column += 1

            channel_sheet.cell(row=2, column=1).value = primary_page

            incremental_row = 3
            for entry in self.validation_data:
                channel_sheet.cell(row=incremental_row, column=1).value = entry["packaged_service_id"]
                channel_sheet.cell(row=incremental_row, column=2).value = entry["stationId"]
                channel_sheet.cell(row=incremental_row, column=3).value = entry["call_sign"]

                channel_sheet.cell(row=incremental_row, column=4).value = entry["channelName"]
                channel_sheet.cell(row=incremental_row, column=5).value = entry["packaged_service_description"]

                channel_sheet.cell(row=incremental_row, column=7).value = entry["lcn"]
                channel_sheet.cell(row=incremental_row, column=8).value = entry["url"]
                channel_sheet.cell(row=incremental_row, column=11).value = "hlsTransportStream"

                incremental_row += 1

        for msoServiceID in self.list_msoservice_IDs[1:]:
            copy_page = self.workbook.copy_worksheet(channel_sheet)
            copy_page.title = msoServiceID
            copy_page["A2"] = msoServiceID


    def save_file(self):
        try:
            self.excel_file_path = f"Xumo_US_120_file{self.xumo_partner_id}_file.xlsx"
            self.workbook.save(self.excel_file_path)
        except PermissionError:
            print("Close the excel file before saving it")


if __name__ == "__main__":
    #################
    list_of_operators = [["bluestream","1007704","bluestream_IP_LINEAR_LINEAR_SERVICE_Export_20250626-080457.xlsx"],["armstrong","1007769","armstrong_IP_LINEAR_LINEAR_SERVICE_Export_20250626-080726.xlsx"],["midco","1007163","midco_IP_LINEAR_LINEAR_SERVICE_Export_20250626-081004.xlsx"]]
    # confluence page
    xumo_source_file = "xumo_120.csv"
    xumo_ca_packaged_service_id = 99000000
    xumo_packaged_service_id = 99000

    for operator in list_of_operators:
        xumo_partner_id = operator[1]
        # operator SC file
        operator_file =  operator[2]

        ###################
        list_msoservice_IDs = read_mso_localities(xumo_packaged_service_id, operator_file)
        file_processor = Xumo_processor(xumo_source_file, xumo_partner_id, list_msoservice_IDs,xumo_packaged_service_id, operator_file)

        file_processor.xumo_csv_to_data()
        file_processor.open_mso_file_data()
        file_processor.compare_and_filter()
        file_processor.write_channel_page()
        file_processor.write_msoSiteID_pages()
        file_processor.save_file()