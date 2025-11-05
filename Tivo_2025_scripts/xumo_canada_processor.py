import csv
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import re
import pandas
from openpyxl.workbook import Workbook


class XumoCA_processor():
    def __init__(self,solution_source_file,xumo_partner_id,list_msoservice_IDs,channel_start_range,packaged_service_id,mso_all_data_file,folder,file_name,blacklisted_stations):
        self.workbook = Workbook()
        self.solution_source_file = solution_source_file
        self.channel_start_range = channel_start_range
        self.list_msoservice_IDs = list_msoservice_IDs
        self.xumo_partner_id = xumo_partner_id
        self.mso_all_data_file = mso_all_data_file
        self.packaged_service_id = packaged_service_id
        self.folder = folder
        self.file_name = file_name
        self.blacklisted_stations = blacklisted_stations


    # Regex strip pattern
    pattern = re.compile(r';"(.+)"')

    def xumo_csv_to_data(self):
        first_col = self.solution_source_file.columns[0]
        filtered_df = self.solution_source_file[self.solution_source_file[first_col].astype(str).str.isdigit()]
        filtered_df['Partner Station Id'] = filtered_df['Partner Station Id'].astype('Int64')
        channel_data = filtered_df.values.tolist()

        list_confluence_channels = [{"start_date": "01/01/1970 12:00:01 AM", "end_date": "12/31/2099 11:59:59 PM", "linear_channel_nr": None,"packaged_service_id": None,"call_sign": entry[4],
                                     "channelName": entry[1], "stationId": "epgProvider:st." + str(entry[3]),"url": entry[6], "channel_description": entry[5]} for entry in channel_data]


        # Adding LCN to the linear channels
        for channel in list_confluence_channels:
            channel["linear_channel_nr"] = self.channel_start_range
            self.channel_start_range += 1

        # Removing empty elements from the list
        pattern_stationid = re.compile(r'epgProvider:st.(\d+)')
        self.list_confluence_channels_stripped = [item for item in list_confluence_channels if len(pattern_stationid.findall(item["stationId"])) >= 1]

        # Removing the =URL formatting if found at channel level
        # Regex strip pattern
        pattern = re.compile(r';"(.+)"')
        for channel in self.list_confluence_channels_stripped:
            url_search = pattern.findall(channel["url"])
            if url_search:
                channel["url"] = url_search[0]

            else:
                pass

        # Only custom rule setup as Amanda keeps leaving the channel as it is
        for channel in self.list_confluence_channels_stripped:
            if channel["stationId"] == "epgProvider:st.10427289051":
                channel["stationId"] = "epgProvider:st.446201192"

        return self.list_confluence_channels_stripped

    def open_mso_file_data(self):
        packaged_service_id = self.packaged_service_id
        data = pandas.read_excel(self.mso_all_data_file)
        all_data = data.loc[:,["Packaged Service Id", "Partner Station Id", "Station Name", "Availability Window Start","Availability Window End"]]
        y = all_data.query(f'{packaged_service_id} <= `Packaged Service Id`  < {packaged_service_id + 400}')
        self.pckd_serv_id = list(y["Packaged Service Id"])
        prtn_st_id = list(y["Partner Station Id"])
        channel_st_id = list(y["Station Name"])
        start_date = list(y["Availability Window Start"])
        end_date = list(y["Availability Window End"])
        self.validation_data = [
            {"channelName": channel_st_id[i], "stationId": prtn_st_id[i], "packaged_service_id": self.pckd_serv_id[i], "start_date": start_date[i], "end_date": end_date[i]} for i in range(len(prtn_st_id))]

        return self.validation_data, self.pckd_serv_id

    def compare_and_filter(self):
        #self.list_confluence_channels_stripped, self.validation_data, self.pckd_serv_ids, packaged_service_id
        for conf_channel in self.list_confluence_channels_stripped:
            for operator_channel in self.validation_data:
                if conf_channel["stationId"] == operator_channel["stationId"]:
                    conf_channel["packaged_service_id"] = operator_channel["packaged_service_id"]
                    conf_channel["start_date"] = operator_channel["start_date"]
                    conf_channel["end_date"] = operator_channel["end_date"]

        for conf_channel in self.list_confluence_channels_stripped:
            if conf_channel["packaged_service_id"] == None:
                check = True
                while check:
                    if self.packaged_service_id not in self.pckd_serv_id:
                        conf_channel["packaged_service_id"] = self.packaged_service_id
                        self.packaged_service_id += 1
                        check = False
                    else:
                        self.packaged_service_id += 1

        return self.list_confluence_channels_stripped

    def write_channel_page(self):
        channel_sheet = self.workbook["Sheet"]
        channel_sheet.title = "Channel"

        header_data = ["Packaged Service Id", "Partner Station Id", "Call Sign", "Station Name","Packaged Service Description", "Jump Channel Type", "Availability Window Start", \
                       "Availability Window End", "Linear Service Type", "DVB Linkage Id", "Linear Provider Partner Id","Recording Provider Partner Id", \
                       "Partner Channel Id", "Video Resolution", "Prevent EAS Interruption", "Channel Change Id", \
                       "Playback URI Channel Id", "IP ABR URL", "DRM Type", "Transport Encoding Type", "Device Type","SOCU Base URL", "Application Id", "Channel Name", \
                       "Channel Description", "Guide Cell Title", "Logo Partner Id", "DRM Content ID", "Multi View"]

        incremental_column = 1

        for value in range(0, len(header_data)):
            channel_sheet.cell(row=1, column=incremental_column).value = header_data[value]
            channel_sheet.cell(row=1, column=incremental_column).fill = PatternFill(start_color="00CCFF",
                                                                                    end_color="00CCFF",
                                                                                    fill_type="solid")
            channel_sheet.column_dimensions[get_column_letter(incremental_column)].width = 25
            incremental_column += 1

        incremental_row = 2
        for entry in self.list_confluence_channels_stripped:
            operator_name = self.file_name.split("-")[0]
            if (operator_name in self.blacklisted_stations) and (entry["stationId"] in self.blacklisted_stations[operator_name]):
                pass

            else:
                channel_sheet.cell(row=incremental_row, column=1).value = entry["packaged_service_id"]
                channel_sheet.cell(row=incremental_row, column=2).value = entry["stationId"]
                channel_sheet.cell(row=incremental_row, column=3).value = entry["call_sign"]

                channel_sheet.cell(row=incremental_row, column=4).value = entry["channelName"]
                channel_sheet.cell(row=incremental_row, column=5).value = entry["channel_description"]

                channel_sheet.cell(row=incremental_row, column=7).value = entry["start_date"]
                channel_sheet.cell(row=incremental_row, column=8).value = entry["end_date"]
                channel_sheet.cell(row=incremental_row, column=9).value = "linear"
                channel_sheet.cell(row=incremental_row, column=11).value = f"tivo:pt.{self.xumo_partner_id}"

                channel_sheet.cell(row=incremental_row, column=15).value = "false"

                channel_sheet.cell(row=incremental_row, column=18).value = entry["url"]
                channel_sheet.cell(row=incremental_row, column=20).value = "hlsTransportStream"

                incremental_row += 1
        return channel_sheet

    def write_msoSiteID_pages(self):
        #self.confluence_file, list_msoservice_IDs
        page_columns = ["Packaged Service Id", "Partner Station Id", "Call Sign", "Station Name","Packaged Service Description", "Partner Channel Id", \
                        "Logical Channel Number", "IP ABR URL", "Channel Change Id", "Playback URI Channel Id", "Transport Encoding Type", "Preferred Transport", "Device Type", "DRM Type", \
                        "SOCU Base URL", "Application Id", "Channel Name", "Channel Description", "Guide Cell Title","Logo Partner Id" \
                        ]
        primary_page = self.list_msoservice_IDs[0]
        self.workbook.create_sheet(primary_page)
        channel_sheet = self.workbook[primary_page]

        incremental_column = 1
        for data in range(len(page_columns)):
            channel_sheet.cell(row=1, column=incremental_column).value = page_columns[data]
            channel_sheet.cell(row=1, column=incremental_column).fill = PatternFill(start_color="00CCFF",
                                                                                    end_color="00CCFF",
                                                                                    fill_type="solid")
            channel_sheet.column_dimensions[get_column_letter(incremental_column)].width = 25
            incremental_column += 1

            channel_sheet.cell(row=2, column=1).value = primary_page

            incremental_row = 3
            for entry in self.list_confluence_channels_stripped:
                operator_name = self.file_name.split("-")[0]
                if (operator_name in self.blacklisted_stations) and (entry["stationId"] in self.blacklisted_stations[operator_name]):
                    pass

                else:
                    channel_sheet.cell(row=incremental_row, column=1).value = entry["packaged_service_id"]
                    channel_sheet.cell(row=incremental_row, column=2).value = entry["stationId"]
                    channel_sheet.cell(row=incremental_row, column=3).value = entry["call_sign"]

                    channel_sheet.cell(row=incremental_row, column=4).value = entry["channelName"]
                    channel_sheet.cell(row=incremental_row, column=5).value = entry["channel_description"]

                    channel_sheet.cell(row=incremental_row, column=7).value = entry["linear_channel_nr"]
                    channel_sheet.cell(row=incremental_row, column=8).value = entry["url"]
                    channel_sheet.cell(row=incremental_row, column=11).value = "hlsTransportStream"

                    incremental_row += 1

        for msoServiceID in self.list_msoservice_IDs[1:]:
            copy_page = self.workbook.copy_worksheet(channel_sheet)
            copy_page.title = msoServiceID
            copy_page["A2"] = msoServiceID


    def save_file(self):
        try:
            self.excel_file_path = f"{self.folder}/Xumo_CA_{self.file_name}_file.xlsx"
            self.workbook.save(self.excel_file_path)
        except PermissionError:
            print("Close the excel file before saving it")
