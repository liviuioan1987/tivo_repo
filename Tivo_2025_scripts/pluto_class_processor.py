import csv
import sys
import pandas
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

class Pluto_processor:
    def __init__(self,solution_source_file,pluto_partner_id,list_msoservice_IDs,channel_start_range,folder,file_name):
        self.workbook = openpyxl.Workbook()
        self.solution_source_file = solution_source_file
        self.pluto_partner_id = pluto_partner_id
        self.list_msoservice_IDs = list_msoservice_IDs
        self.channel_start_range = channel_start_range
        self.folder = folder
        self.file_name = file_name


    def pluto_csv_to_data(self):
        first_col = self.solution_source_file.columns[0]
        filtered_df = self.solution_source_file[self.solution_source_file[first_col].astype(str).str.isdigit()]
        filtered_df['Partner Station Id'] = filtered_df['Partner Station Id'].astype('Int64')
        channel_data = filtered_df.values.tolist()
        self.page_row = channel_data

        return self.page_row

    def write_channel_page(self):
        channel_sheet = self.workbook["Sheet"]
        channel_sheet.title = "Channel"

        header_data = ["Packaged Service Id", "Partner Station Id", "Call Sign", "Station Name","Packaged Service Description", "Jump Channel Type", "Availability Window Start", \
                       "Availability Window End", "Linear Service Type", "DVB Linkage Id", "Linear Provider Partner Id","Recording Provider Partner Id", \
                       "Partner Channel Id", "Video Resolution", "Prevent EAS Interruption", "Channel Change Id", \
                       "Playback URI Channel Id", "IP ABR URL", "DRM Type", "Transport Encoding Type", "Device Type","SOCU Base URL", "Application Id", "Channel Name", \
                       "Channel Description", "Guide Cell Title", "Logo Partner Id", "DRM Content ID", "Multi View"]

        incremental_column = 1
        for value in range(0, len(header_data)):
            channel_sheet.cell(row=1, column=incremental_column).value = header_data[value]
            channel_sheet.cell(row=1, column=incremental_column).fill = PatternFill(start_color="00CCFF", end_color="00CCFF",fill_type="solid")
            channel_sheet.column_dimensions[get_column_letter(incremental_column)].width = 25
            incremental_column += 1

        incremental_row = 2
        for row  in range(0, len(self.page_row)):
            if (isinstance(self.page_row[row][3],int)):
                channel_sheet.cell(row=incremental_row, column=1).value = self.page_row[row ][3]
                channel_sheet.cell(row=incremental_row, column=2).value = f"epgProvider:st.{self.page_row[row ][3]}"
                channel_sheet.cell(row=incremental_row, column=3).value = self.page_row[row][4]

                channel_sheet.cell(row=incremental_row, column=4).value = self.page_row[row ][1]
                channel_sheet.cell(row=incremental_row, column=5).value = self.page_row[row ][5]

                channel_sheet.cell(row=incremental_row, column=7).value = "01/01/1970 12:00:01 AM"
                channel_sheet.cell(row=incremental_row, column=8).value = "12/31/2099 11:59:59 PM"

                channel_sheet.cell(row=incremental_row, column=9).value = "appLinear"
                channel_sheet.cell(row=incremental_row, column=11).value = f"tivo:pt.{self.pluto_partner_id}"

                #channel_sheet.cell(row=incremental_row, column=15).value = "false"
                #channel_sheet.cell(row=incremental_row, column=16).value = "[]"

                incremental_row += 1
            else:
                pass

        return channel_sheet

    def write_msoSiteID_pages(self):
        page_columns = ["Packaged Service Id", "Partner Station Id", "Call Sign", "Station Name","Packaged Service Description", "Partner Channel Id", \
                        "Logical Channel Number", "IP ABR URL", "Channel Change Id", "Playback URI Channel Id", "Preferred Transport", "Device Type", "DRM Type","Transport Encoding Type", \
                        "SOCU Base URL", "Application Id", "Channel Name", "Channel Description", "Guide Cell Title","Logo Partner Id" \
                        ]

        for msoServiceID in self.list_msoservice_IDs:
            channel_nr = self.channel_start_range
            self.workbook.create_sheet(msoServiceID)
            channel_sheet = self.workbook[msoServiceID]

            incremental_column = 1
            for data in range(len(page_columns)):
                channel_sheet.cell(row=1, column=incremental_column).value = page_columns[data]
                channel_sheet.cell(row=1, column=incremental_column).fill = PatternFill(start_color="00CCFF", end_color="00CCFF",fill_type="solid")
                channel_sheet.column_dimensions[get_column_letter(incremental_column)].width = 25
                incremental_column += 1

            channel_sheet.cell(row=2, column=1).value = msoServiceID

            incremental_row = 3
            for row in range(0, len(self.page_row)):
                if (isinstance(self.page_row[row][3], int)):
                    channel_sheet.cell(row=incremental_row, column=1).value = self.page_row[row][3]
                    channel_sheet.cell(row=incremental_row, column=2).value = f"epgProvider:st.{self.page_row[row][3]}"
                    channel_sheet.cell(row=incremental_row, column=3).value = self.page_row[row][4]

                    channel_sheet.cell(row=incremental_row, column=4).value = self.page_row[row][1]
                    channel_sheet.cell(row=incremental_row, column=5).value = self.page_row[row][5]
                    channel_sheet.cell(row=incremental_row, column=7).value = channel_nr

                    channel_sheet.cell(row=incremental_row, column=9).value = self.page_row[row][6]
                    channel_sheet.cell(row=incremental_row, column=10).value = self.page_row[row][7]

                    incremental_row += 1
                    channel_nr += 1
                else:
                    channel_nr += 1

    def save_file(self):
        try:
            self.excel_file_path = f"{self.folder}/Pluto_US_{self.file_name}_file.xlsx"
            self.workbook.save(self.excel_file_path)
        except PermissionError:
            print("Close the excel file before saving it")