import pandas as pd
import os
import glob
from datetime import date
import copy
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook

class MediaOps_linpub():
    def __init__(self, mediaops_jumchannels_file):
        self.mops_jumpchannels = mediaops_jumchannels_file

    def extract_data_from_template_file(self):
        self.global_values = {}
        self.operator_files_values = {}

        mops_all_tabs = pd.ExcelFile(self.mops_jumpchannels)
        mops_sheet_names = mops_all_tabs.sheet_names

        for sheet_name in mops_sheet_names:
            df = pd.read_excel(self.mops_jumpchannels, sheet_name=sheet_name)

            if sheet_name == "Global_Values":
                result = {}
                for _, row in df.iterrows():
                    app_name = row["JumpApp name"]
                    app_id = row["ApplicationID"]

                    # Split Device Type into a list, strip spaces
                    device_types = [d.strip() for d in str(row["Device Type"]).split(",") if
                                    pd.notna(row["Device Type"])]

                    # Build entry (keeping other columns too if needed)
                    result[app_name] = (app_id, device_types, row.get("Description", None))

            self.global_values = dict(result)

            if sheet_name == "Operator_Values":
                # Forward-fill missing Operator Name values
                df["Operator Name"] = df["Operator Name"].ffill()

                # Build dictionary
                result = defaultdict(list)

                for _, row in df.iterrows():
                    operator = row["Operator Name"]
                    entry = {
                        "Vod app name": row["Vod app name"],
                        "ApplicationId": row["ApplicationId"],
                        "Source file": row["Source file"] if not pd.isna(row["Source file"]) else ""
                    }
                    result[operator].append(entry)

                # Convert back to normal dict
                self.operator_files_values= dict(result)

        return self.global_values, self.operator_files_values

    def update_excel_file_values(self, operator_total_values, operator_file_list, operator_name):
        operator_file_data = operator_file_list[0]["Source file"]
        operator_file_location = f"{mso_all_data_folder}/{operator_file_data}"

        temp_data = {}
        for value in range(0, len(operator_file_list)):

            if isinstance(operator_file_list[value]['ApplicationId'],str):
                temp_data[operator_file_list[value]['Vod app name']] = operator_file_list[value]['ApplicationId']
                operator_total_values.update(temp_data)

                #astound custom case
                if operator_name == "Astound":
                    self.astound_key = list(temp_data)
            else:
                pass
        try:
            # Load the workbook with formatting preserved
            wb = load_workbook(operator_file_location)

            # Work on each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Identify column indexes by header row (assuming first row is headers)
                headers = {cell.value: cell.column for cell in ws[1]}

                channelName_col = headers.get("Channel Name")
                callSign_col = headers.get("Call Sign")
                packagedServiceDescription_col = headers.get("Packaged Service Description")
                application_ID_col = headers.get("Application Id")
                logo_partner_ID_col = headers.get("Logo Partner Id")
                device_type_col = headers.get("Device Type")
                channel_description_col = headers.get("Channel Description")

                # Iterate rows starting from row 2
                for row in ws.iter_rows(min_row=2):
                    code = row[channelName_col - 1].value
                    if isinstance(code, str):
                        source = (
                            code.replace("R and B", "R&B")
                            .replace("60s and '70s", "60s &'70s")
                            .replace("Alt and Rock", "Alt & Rock")
                            .replace("'70s and '80s", "'70s & '80s")
                            .replace("Singers and Swing", "Singers & Swing")
                            .replace("Pop and Country", "Pop & Country")
                        )
                        row[channelName_col - 1].value = source

                        if source in operator_total_values:
                            row[callSign_col - 1].value = source
                            row[packagedServiceDescription_col - 1].value = source

                            value = operator_total_values[source]
                            if isinstance(value, str):
                                row[application_ID_col - 1].value = value
                                row[logo_partner_ID_col - 1].value = value
                                if operator_name == "Astound":
                                    if row[channelName_col - 1].value in self.astound_key:
                                        row[channel_description_col - 1].value = f"{row[channelName_col - 1].value}."
                            elif isinstance(value, tuple):
                                row[application_ID_col - 1].value = value[0]
                                row[logo_partner_ID_col - 1].value = value[0]
                                row[device_type_col - 1].value = ",".join(value[1])
                                if value[2] and str(value[2]) != 'nan':
                                    row[channel_description_col - 1].value = value[2]

            # Save new file (formatting preserved)
            today = str(date.today())
            directory_text= "UPDATED_files"
            directory = directory_text + "_" + today

            if not os.path.exists(directory):
                os.makedirs(directory)
                print(directory, " Folder created")

            output_file = f"{directory}/Updated_{today}_{operator_name}.xlsx"
            wb.save(output_file)
            print(f"✅ Created {output_file}")

        except Exception as err:
            print(f"❌ Error: {err}")


if __name__ == "__main__":
    ## Mediaops template file processing
    mediaops_file = "mediaops_source_file.xlsx"
    mediaops_source_file_processor = MediaOps_linpub(mediaops_file)
    operator_global_values, operator_specific_values = mediaops_source_file_processor.extract_data_from_template_file()

    # Folder location where the all data files for operators are uploaded
    # iterate each file and apply the updates
    mso_all_data_folder = "Bulk_operator_files"
    for operator_name,operator_file_list in operator_specific_values.items():
        operator_total_values = copy.deepcopy(operator_global_values)
        mediaops_source_file_processor.update_excel_file_values(operator_total_values, operator_file_list, operator_name)



