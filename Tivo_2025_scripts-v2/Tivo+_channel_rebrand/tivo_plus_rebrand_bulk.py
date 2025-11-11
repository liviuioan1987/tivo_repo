import os
from openpyxl import load_workbook
from datetime import date
# 1. Dictionaries for replacement
old_values = {
    "Partner_Station_Id": {"column_name":"Partner Station Id","old_value":""},
    "Station_Name": {"column_name":"Station Name","old_value":""},
    "Packaged_Service_Description": {"column_name":"Packaged Service Description","old_value":""},
    "Channel_Change_Id": {"column_name":"Channel Change Id","old_value":""},
    "Playback_URI_Channel_Id": {"column_name":"Playback URI Channel Id","old_value":""},
    "IP_ABR_URL": {"column_name":"IP ABR URL","old_value":"tes13223"}
}
new_values = {
    "Partner_Station_Id": {"column_name":"Partner Station Id","new_value":"epgProvider:st.new1"},
    "Station_Name": {"column_name":"Station Name","new_value":"CNN new"},
    "Packaged_Service_Description": {"column_name":"Packaged Service Description","new_value":"CNN new2"},
    "Channel_Change_Id": {"column_name":"Channel Change Id","new_value":"new3"},
    "Playback_URI_Channel_Id": {"column_name":"Playback URI Channel Id","new_value":"cnn-new4"},
    "IP_ABR_URL": {"column_name": "IP ABR URL", "new_value": "new5664"}
}

# 2. Locate Excel files in source folder
source_folder = "Tivo+_to_update"  # change as needed
files = [f for f in os.listdir(source_folder) if f.endswith('.xlsx')]

directory = "Updated_files_" + str(date.today())

if not os.path.exists(directory):
    os.makedirs(directory)
    print(directory, "Folder created")

for filename in files:
    file_path = os.path.join(source_folder, filename)
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Get header row (assumed first row)
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        header_map = {cell.value: cell.column for cell in header_row}

        for key in old_values.keys():
            old_val = old_values[key].get("old_value")
            # Skip if old_value is empty
            if not old_val:
                continue
            col_name = old_values[key]["column_name"]
            new_val = new_values[key]["new_value"]
            col_idx = header_map.get(col_name)
            if col_idx is not None:
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    cell = row[0]
                    if cell.value == old_val:
                        cell.value = new_val

    new_filename = f"Update_{filename}"
    out_path = os.path.join(directory, new_filename)
    wb.save(out_path)
    print(f"{filename} updated and saved as {new_filename} in {directory}")