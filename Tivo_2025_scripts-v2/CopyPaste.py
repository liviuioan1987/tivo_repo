msoServiceIDs_list = """Alexandria_ip
Amherst_ip
Antigonish_ip
Aylesford_ip
Aylmer_ip
Baccalieu_ip
Barrington_ip
Bashaw_ip
Bay_Cable_ip
Bedford_Sackville_ip
Blind_river_ip
Blockhouse_ip
Bluewater_ip
Bonnyville_ip
Bourget_ip
Cambray_ip
Campbellford_ip
Camrose_ip
Casselman_ip
Charlottetown_ip
Chetwynd_ip
Clare_ip
Clarenville_ip
Coast_ip
Cochrane_ip
Coldwater_ip
Dartmouth_ip
Delta_ip
Diamond_Valley_ip
Digby_ip
Elliot_Lake_ip
Espanola_ip
Forest_ip
Glencoe_ip
Golden_ip
Grand_Prairie_ip
Grimshaw_ip
Halifax_ip
Hanover_ip
Happy_Valley_ip
Harbour_Breton_ip
Harbour_Grace_ip
Irricana_ip
Island_Rural_ip
Kapuskasing_ip
Kirkland_Lake_ip
Langdon_ip
Lewisporte_ip
Limoges_ip
Liverpool_ip
Markdale_ip
Marystown_ip
Marysvale_ip
Mayerthorpe_ip
Milford_ip
Milltown_ip
New_Germany_ip
New_Glasgow_ip
New_Liskeard_ip
New_Minas_ip
Oliver_ip
Oxford_ip
Parrsboro_ip
Peace_River_ip
Picton_ip
Port_Elgin_ip
Port_Hawkesbury_ip
Porters_Lake_ip
Princeton_ip
Pubnico_ip
Redwood_Meadows_ip
Sackville_NB_ip
Shelburne_ip
Simcoe_ip
Slave_Lake_ip
Springdale_ip
St_Lawrence_Shores_ip
Stephenville_ip
Sturgeon_Falls_ip
Sudbury_ip
Summerside_Rebuilt_ip
Sydney_ip
Three_Hills_ip
Timmins_ip
Truro_ip
Wainwright_ip
Wetaskiwin_ip
Whitecourt_ip
Windsor_ip
Yarmouth_ip"""

import re
from openpyxl import load_workbook


#DeliveryFE Input data here#
mso_all_data_file = "eastlink_template_copypaste.xlsx"
output_file = "/Users/liviu.gherasim/Downloads/eastlink_tivoplus_alllocalities_prod_Plex_CA.xlsx"


msoServiceIDs= msoServiceIDs_list.replace("\n", ",").replace(" ", "").split(",")
#print(len(msoServiceIDs))

## OR needed create a list of MSO service ID's from what the MSO has in their export
#mso_export_file ="C:\\Users\lgherasim\Downloads\IP_LINEAR_LINEAR_SERVICE_Export_20231026-085652.xlsx"

def funct_msoServiceIDs(mso_export_file):

    data = load_workbook(mso_export_file)
    raw_list_of_msoids = data.sheetnames
    list_of_msoids = list(map(lambda word: re.sub(r"\d{1,4}\-","",word),raw_list_of_msoids))
    list_of_msoids.pop(0)
    #print(list_of_msoids)
    return list_of_msoids



class Excel:
    def __init__(self,input_file,msoservicelist):
        self.input_file = input_file
        self.msoservicelist = msoservicelist


    def load_file(self):
        self.data = load_workbook(filename=self.input_file)
        self.template_sheet = self.data.sheetnames[1]
        self.mso_worksheet = self.data[self.template_sheet]


    def write_new_file(self):
        if str(self.mso_worksheet["A2"].value) in self.msoservicelist:
            self.msoservicelist.remove(str(self.mso_worksheet["A2"].value))

        for msoServiceID in self.msoservicelist:
            self.copy_page = self.data.copy_worksheet(self.mso_worksheet)
            self.copy_page.title = msoServiceID
            self.copy_page["A2"] = msoServiceID

    """    def delete_template_sheet(self):
            self.data.remove(self.mso_worksheet)
            #print(self.data.sheetnames)"""

    def save_file(self,output_file):
        print("Created MSOLocalities tabs:", len(self.data.sheetnames))
        self.data.save(output_file)

#msoServiceIDs = funct_msoServiceIDs(mso_export_file)

report = Excel(mso_all_data_file,msoServiceIDs)
report.load_file()
report.write_new_file()
#report.delete_template_sheet()
report.save_file(output_file)





