import sys
import os
import time
from datetime import date
import openpyxl
import pandas
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import re

from pluto_class_processor import Pluto_processor
from pluto_canada_processor import PlutoCA_processor
from plex_class_processor import Plex_processor
from plex_canada_class_processor import PlexCA_processor
from xumo_class_processor import Xumo_processor
from xumo_canada_processor import XumoCA_processor
from  format_blacklisted_channels_list import create_dict_blacklisted

import warnings
warnings.filterwarnings("ignore", message="Workbook contains no default style, apply openpyxl's default")

#Tivo+ Sharepoint filename
sharepoint_file_name = "TiVo Plus IPTV and TMIS Channel Lineups.xlsx"

#Tivo+ start ranges for Plex and Xumo
#Pluto
pluto_ca_packaged_service_id = 97000000
pluto_packaged_service_id = 0
#Plex
plex_ca_packaged_service_id = 98000000
plex_us_packaged_service_id = 0
#Xumo
xumo_ca_packaged_service_id = 99000000
xumo_packaged_service_id = 99000

##Blacklisted stations
blacklist_file = "Stingray Blacklisted Channels - PayTV.xlsx"
blacklist_location = f"TivoPlus_source_files/{blacklist_file}"

#Folder location where the all data files for operators are uploaded
mso_all_data_folder  = "Bulk_operator_files"
#Location of the Excel file that contains the operators data
operators_list = "operators_list.xlsx"

#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
operators_excel_file = openpyxl.load_workbook(operators_list, data_only=True)
production_sheet = operators_excel_file["Production"]
staging_sheet = operators_excel_file["Staging"]

#Production section
production_generator = production_sheet.values
data = next(production_generator)
#Staging section
staging_generator = staging_sheet.values
data = next(staging_generator)

#Variables for folders creation + processing of environments Excel Tabs
production = "PROD_operators"
staging  = "STG_operators"

stg_prod_folders = [production,staging]
sheet_names = [production_generator, staging_generator]
env_generator_save_folder = list(zip(sheet_names,stg_prod_folders))
#print(env_generator_save_folder)

def sharepoint_source_to_tivoplus_data(file_path):
    file = pandas.ExcelFile(file_path)

    us_files  = file.parse("IPTV US")
    ca_files = file.parse("IPTV CA")

    # US files
    #Xumo file generator
    us_filter_xumo = ["Direct","XUMO"]
    #Xumo pre-frumos
    us_xumo_pre_frumos = us_files.loc[:, ["Channel Number", "Channel Name", "Provider", "Partner Station Id", "Call Sign","Packaged Service Description","XUMO playURL (Pre-Frumos 1.20)"]]
    us_xumo_pre_frumos = us_xumo_pre_frumos.loc[us_xumo_pre_frumos["Provider"].isin(us_filter_xumo)]

    #Xumo 1.20
    us_xumo_1_20 = us_files.loc[:,["Channel Number", "Channel Name", "Provider", "Partner Station Id", "Call Sign","Packaged Service Description", "XUMO playURL (Frumos 1.20)"]]
    us_xumo_1_20 = us_xumo_1_20.loc[us_xumo_1_20["Provider"].isin(us_filter_xumo)]

    #Plex US
    us_filter_plex = ["Plex"]
    us_plex = us_files.loc[:,["Channel Number", "Channel Name", "Provider", "Partner Station Id", "Call Sign","Packaged Service Description", "channelChangeId", "playbackUrlChannelId"]]
    us_plex = us_plex.loc[us_plex["Provider"].isin(us_filter_plex)]

    #Pluto US
    us_filter_pluto = ["Pluto TV"]
    us_pluto= us_files.loc[:,["Channel Number", "Channel Name", "Provider", "Partner Station Id", "Call Sign","Packaged Service Description", "channelChangeId", "playbackUrlChannelId"]]
    us_pluto = us_pluto.loc[us_pluto["Provider"].isin(us_filter_pluto)]

    #Canada
    #Xumo file generator
    ca_filter_xumo = ["Direct","XUMO"]
    ca_xumo = ca_files.loc[:,["Channel Number", "Channel Name", "Provider", "Partner Station Id", "Call Sign","Packaged Service Description", "XUMO playURL (Frumos 1.20)"]]
    ca_xumo = ca_xumo.loc[ca_xumo["Provider"].isin(ca_filter_xumo)]

    #Plex CA
    ca_filter_plex = ["Plex"]
    ca_plex = ca_files.loc[:,["Channel Number", "Channel Name", "Provider", "Partner Station Id", "Call Sign","Packaged Service Description", "channelChangeId", "playbackUrlChannelId"]]
    ca_plex = ca_plex.loc[ca_plex["Provider"].isin(ca_filter_plex)]


    return us_pluto,us_xumo_pre_frumos,us_xumo_1_20,us_plex,ca_xumo,ca_plex


def generate_properties_files_in_bulk(env_generator):
        operators_generate_files = []
        for row in env_generator:
            operator_save_file = row[0]
            operator_partner_id = row[1]
            pluto_start_range = row[3]
            xumo_start_range = row[4]
            plex_start_range = row[5]
            operator_file = f"{mso_all_data_folder}/{row[7]}"

            ##Hotwire initial custom rule. To be removed after onboarding
            ##Required only on initial onboardings where MSO Localities are a long string value. After the value can be removed
            new_operator_name = "Hotwire-Production"
            if new_operator_name== operator_save_file:
                operator_msolocalities = "009,010,011,014,015,016,017,018,019,022,025,027,028,030,032,033,034,035,036,037,038,039,040,041,042,043,044,046,048,066,068,069,999"
            else:
                operator_msolocalities = row[8]

            operator_canada_mso_localities = row[9]
            xumo_120_parameter = row[10]
            station_policy = row[11]
            linear_packages = row[12]
            linear_packages_file = f"{mso_all_data_folder}/{row[13]}"
            operators_generate_files.append([operator_save_file,operator_partner_id,pluto_start_range,xumo_start_range,plex_start_range, operator_file,operator_msolocalities,operator_canada_mso_localities, xumo_120_parameter,station_policy,linear_packages,linear_packages_file])
        return operators_generate_files

def read_mso_localities(source,operator_name, packaged_service_id,operator_file,partner_id, pluto="yes"):
    tivo_plus_localities = []
    file = pandas.ExcelFile(operator_file)
    all_localities = file.sheet_names

    reading_page = file.parse("Channel")
    if packaged_service_id == 0:
        operator_channel_page_query = reading_page.loc[reading_page["Linear Provider Partner Id"] == f'tivo:pt.{partner_id}']
    else:
        operator_channel_page_query = reading_page.query(f'{packaged_service_id} <= `Packaged Service Id`  < {packaged_service_id + 800}')
    operator_channel_page = operator_channel_page_query["Packaged Service Id"].size

    for locality in all_localities[1:]:
        reading_pages  = file.parse(locality)
        try:
            reading_pages["Packaged Service Id"] = pandas.to_numeric(reading_pages["Packaged Service Id"],downcast='integer', errors='coerce')

            if packaged_service_id == 0 and pluto == "yes":
                reading_pages["Partner Station Id"] = reading_pages["Partner Station Id"].str.split('.').str[1]
                reading_pages["Partner Station Id"] = pandas.to_numeric(reading_pages["Partner Station Id"],downcast='integer', errors='coerce')
                operator_locality_page_query = reading_pages.loc[(reading_pages["Packaged Service Id"] == reading_pages["Partner Station Id"]) & (reading_pages["Channel Change Id"].str.contains("5e20b730f2f8d5003d739db7") == False )]
                #print(operator_locality_page_query)

            elif packaged_service_id == 0 and pluto == "no":
                reading_pages["Partner Station Id"] = reading_pages["Partner Station Id"].str.split('.').str[1]
                reading_pages["Partner Station Id"] = pandas.to_numeric(reading_pages["Partner Station Id"],downcast='integer', errors='coerce')
                operator_locality_page_query = reading_pages.loc[(reading_pages["Packaged Service Id"] == reading_pages["Partner Station Id"]) & (reading_pages["Channel Change Id"].str.contains("5e20b730f2f8d5003d739db7") == True)]
                #print(operator_locality_page_query)

            elif packaged_service_id != 0:
                operator_locality_page_query = reading_pages.query(f'{packaged_service_id} <= `Packaged Service Id`  < {packaged_service_id + 800}')

            operator_locality_page = operator_locality_page_query["Packaged Service Id"].size

            if operator_locality_page > 2:
                if operator_locality_page != operator_channel_page:
                    print(f"{source}: For {operator_name}: In locality {locality} not all linear channels were present, there are {operator_locality_page} and were expected {operator_channel_page}")
                    pass
                if "-" in locality:
                    locality = locality.split("-")
                    tivo_plus_localities.append(locality[1])
                else:
                    tivo_plus_localities.append(locality)
        except:
            pass
    return tivo_plus_localities

def station_policy_generator(operator_name,folder,station_policy_file,policy_name):
    if station_policy_file != None:
        workbook = Workbook()
        channel_sheet = workbook["Sheet"]
        channel_sheet.title  = "Station_Policy_Map"
        header_data = ["Call Sign","Station Name","Partner Station Id","nDVR Station Policy Name","Linear Station Policy Name","SOCU Station Policy Name"]
        incremental_column = 1
        for value in range(0, len(header_data)):
            channel_sheet.cell(row=1, column=incremental_column).value = header_data[value]
            channel_sheet.cell(row=1, column=incremental_column).fill = PatternFill(start_color="00CCFF",end_color="00CCFF",fill_type="solid")
            channel_sheet.column_dimensions[get_column_letter(incremental_column)].width = 25
            incremental_column += 1

        policy_data_source = []
        for row in station_policy_file.iter_rows(min_row=2, values_only=True):
                policy_data_source.append([row[2],row[3],row[1]])

        incremental_row = 2
        for entry in policy_data_source:
            channel_sheet.cell(row=incremental_row, column=1).value = entry[0]
            channel_sheet.cell(row=incremental_row, column=2).value = entry[1]
            channel_sheet.cell(row=incremental_row, column=3).value = entry[2]
            channel_sheet.cell(row=incremental_row, column=5).value = policy_name

            incremental_row += 1

        try:
            if "eastlink" in operator_name.lower():
                excel_file_path = f"{folder}/Xumo_CA_{operator_name}_station_policy_file.xlsx"
            else:
                excel_file_path = f"{folder}/Xumo_US_{operator_name}_station_policy_file.xlsx"
            workbook.save(excel_file_path)
        except PermissionError:
            print("Close the Station policy map excel file before saving it")

def linear_packages_generator(operator_name=0,folder=0,
                              pluto_ca_packaged_service_id=0,
                              plex_ca_packaged_service_id=0,
                              xumo_ca_packaged_service_id=0,
                              xumo_packaged_service_id=0,
                              xumo_file=0,pluto_file=0,plex_file=0,
                              operator_linear_packages=0,
                              operator_linear_packages_file=0):

    # Helper: accept either a single workbook or a list of workbooks
    def make_file_list(f):
        if not f:
            return []
        if isinstance(f, list) or isinstance(f, tuple):
            return [x for x in f if x]
        return [f]

    # SPECIAL RULE: CableCo11-Staging
    if operator_name == "CableCo11-Staging":

        # Normalize helper for package title keys
        def normalize_title(t):
            return "".join(ch.lower() for ch in str(t) if ch.isalnum())

        # Try to build a mapping of normalized Package Title -> Package Id from operator_linear_packages_file (if given)
        package_id_map = {}
        try:
            if operator_linear_packages_file:
                df_pk = pandas.read_excel(operator_linear_packages_file)

                # add direct mappings for exact titles
                if "Package Title" in df_pk.columns and "Package Id" in df_pk.columns:
                    for _, r in df_pk.iterrows():
                        key = normalize_title(r["Package Title"])
                        package_id_map[key] = r["Package Id"]

                # coerce Packaged Service Id to numeric for heuristic categorization
                if "Packaged Service Id" in df_pk.columns:
                    df_pk["Packaged Service Id"] = pandas.to_numeric(df_pk["Packaged Service Id"], downcast='integer', errors='coerce')

                    # build counts per Package Title: CA vs US
                    title_stats = {}
                    for _, r in df_pk.dropna(subset=["Package Title"]).iterrows():
                        title = r["Package Title"]
                        pid = r["Packaged Service Id"]
                        title_stats.setdefault(title, {"ca": 0, "us": 0, "example_package_id": None})
                        if pandas.notna(pid):
                            pid_int = int(pid)
                            if (
                                (xumo_ca_packaged_service_id <= pid_int < xumo_ca_packaged_service_id + 800) or
                                (plex_ca_packaged_service_id <= pid_int < plex_ca_packaged_service_id + 800) or
                                (pluto_ca_packaged_service_id <= pid_int < pluto_ca_packaged_service_id + 800)
                            ):
                                title_stats[title]["ca"] += 1
                            else:
                                title_stats[title]["us"] += 1
                        # remember a package id value for this title to use as mapping
                        if title_stats[title]["example_package_id"] is None and "Package Id" in df_pk.columns:
                            title_stats[title]["example_package_id"] = r.get("Package Id")

                    # If any title clearly contains CA ids, map it to TIVOPLUS_CA Package
                    for title, stats in title_stats.items():
                        if stats["ca"] > 0:
                            package_id_map[normalize_title("TIVOPLUS_CA Package")] = stats["example_package_id"]

                    # For US, pick the title with the largest 'us' count (fallback to any)
                    best_us_title = None
                    best_us_count = -1
                    for title, stats in title_stats.items():
                        if stats["us"] > best_us_count:
                            best_us_count = stats["us"]
                            best_us_title = title
                    if best_us_title:
                        package_id_map[normalize_title("TIVOPLUS_US Package")] = title_stats[best_us_title]["example_package_id"]

                # Prefer explicit TIVOPLUS_* package ids if present
                if normalize_title("TIVOPLUS_CA Package") in package_id_map:
                    # keep it
                    pass
                if normalize_title("TIVOPLUS_US Package") in package_id_map:
                    pass

        except Exception as e:
            print("Warning: failed reading operator_linear_packages_file for CableCo11:", e)
            package_id_map = {}

        # If mapping still lacks canonical keys, try fuzzy heuristics as a fallback
        if not any('tivoplus' in k for k in package_id_map.keys()):
            ca_key = None
            us_key = None
            for k in package_id_map.keys():
                if 'ca' in k or 'canada' in k:
                    ca_key = k
                if 'us' in k or 'unitedstates' in k:
                    us_key = k
            for k in package_id_map.keys():
                if ca_key is None and ('tivoplus' in k and 'ca' in k):
                    ca_key = k
                if us_key is None and ('tivoplus' in k and ('us' in k or 'usa' in k)):
                    us_key = k
            if ca_key:
                package_id_map[normalize_title("TIVOPLUS_CA Package")] = package_id_map[ca_key]
            if us_key:
                package_id_map[normalize_title("TIVOPLUS_US Package")] = package_id_map[us_key]

        linear_packages_to_openpyxl = Workbook()
        sheet = linear_packages_to_openpyxl.active
        sheet.title = "Linear_Package"

        header = [
            "Package Id","Package Title","Packaged Service Id",
            "Channel Number","Channel Name",
            "Partner Station Id","Call Sign",
            "Provider","Partner"
        ]
        sheet.append(header)

        for col in range(1, len(header) + 1):
            sheet.cell(row=1, column=col).fill = PatternFill(
                start_color="00CCFF", end_color="00CCFF", fill_type="solid"
            )
            sheet.column_dimensions[get_column_letter(col)].width = 25

        packages = {
            "TIVOPLUS_US Package": [],
            "TIVOPLUS_CA Package": []
        }

        # build list of workbooks to process (support single workbook or list)
        data_files = []
        for name, f in (("XUMO", xumo_file), ("PLUTO", pluto_file), ("PLEX", plex_file)):
            for wf in make_file_list(f):
                data_files.append((name, wf))

        # Diagnostics counters
        totals = {"TIVOPLUS_US Package": 0, "TIVOPLUS_CA Package": 0}

        # For CableCo11: determine packaged-service-id column from header if possible,
        # then coerce that value to int even if it's a numeric string.
        for df_name, data_file in data_files:
            try:
                header_row = next(data_file.iter_rows(min_row=1, max_row=1, values_only=True))
                # stricter detection: prefer exact Packaged Service Id header (avoid matching Description)
                pkg_idx = next((i for i, h in enumerate(header_row) if h and any(k in str(h).lower() for k in ['packaged service id', 'packaged id', 'packagedserviceid'])), 6)
            except StopIteration:
                pkg_idx = 6

            sample_count = 0
            print(f"--- Processing {df_name} file for CableCo11 (rows preview) ---")
            for row in data_file.iter_rows(min_row=2, values_only=True):
                # get raw packaged service id using detected column index
                raw_pkg_id = row[pkg_idx] if len(row) > pkg_idx else None
                packaged_service_id = None
                if raw_pkg_id is not None:
                    # If it's already numeric (int/float), coerce to int
                    if isinstance(raw_pkg_id, (int, float)) and not pandas.isna(raw_pkg_id):
                        try:
                            packaged_service_id = int(raw_pkg_id)
                        except Exception:
                            packaged_service_id = None
                    else:
                        # Try to convert numeric-looking string to int (handles '99000000' or '99000000.0')
                        s = str(raw_pkg_id).strip()
                        # remove common formatting characters
                        s_clean = re.sub(r'[^\d.-]', '', s)
                        try:
                            packaged_service_id = int(s_clean)
                        except Exception:
                            try:
                                packaged_service_id = int(float(s_clean))
                            except Exception:
                                packaged_service_id = None

                # Determine CA vs US using numeric packaged_service_id; treat non-numeric as US by default
                is_ca = False
                if packaged_service_id is not None:
                    if (
                        (xumo_ca_packaged_service_id <= packaged_service_id < xumo_ca_packaged_service_id + 800) or
                        (plex_ca_packaged_service_id <= packaged_service_id < plex_ca_packaged_service_id + 800) or
                        (pluto_ca_packaged_service_id <= packaged_service_id < pluto_ca_packaged_service_id + 800)
                    ):
                        is_ca = True

                package_title = "TIVOPLUS_CA Package" if is_ca else "TIVOPLUS_US Package"
                package_id = package_id_map.get(normalize_title(package_title), "")

                packages[package_title].append([
                    package_id,
                    package_title,
                    row[pkg_idx] if len(row) > pkg_idx else None,
                    row[0] if len(row) > 0 else None,
                    row[1] if len(row) > 1 else None,
                    row[2] if len(row) > 2 else None,
                    row[3] if len(row) > 3 else None,
                    row[4] if len(row) > 4 else None,
                    row[5] if len(row) > 5 else None
                ])

                totals[package_title] += 1

                if sample_count < 8:
                    print(f"{df_name} row sample #{sample_count+1}: raw_pkg_id={raw_pkg_id!r}, parsed={packaged_service_id!r}, is_ca={is_ca}, package_title={package_title}, channel_name={row[1] if len(row)>1 else None}")
                sample_count += 1

            print(f"{df_name} file done: produced {sample_count} rows")

        # Append rows and save
        for package_title, rows in packages.items():
            for r in rows:
                sheet.append(r)

        excel_file_path = f"{folder}/{operator_name}_linear_packages_file.xlsx"
        linear_packages_to_openpyxl.save(excel_file_path)

        print("CableCo11 linear packages summary:", totals)
        return

    # Rules for the rest of the operators
    if operator_linear_packages != None and operator_linear_packages_file != None:
        operator_linear_packages_file = pandas.read_excel(operator_linear_packages_file)
        linear_packages = operator_linear_packages.replace("\n", ",").split(",")
        operator_data = operator_linear_packages_file[
            operator_linear_packages_file["Package Title"].isin(linear_packages)
        ]

        linear_packages_list_ids = []
        for linear_package in linear_packages:
            linear_packages_dict_ids = {"package_id": "", "package_title": ""}
            mask = operator_data["Package Title"] == linear_package
            if mask.any():
                linear_packages_dict_ids["package_title"] = linear_package
                linear_packages_dict_ids["package_id"] = operator_data.loc[
                    mask, "Package Id"
                ].iloc[0]
                linear_packages_list_ids.append(linear_packages_dict_ids)

        operator_data = operator_data.copy()
        operator_data["Packaged Service Id"] = pandas.to_numeric(
            operator_data["Packaged Service Id"],
            downcast='integer', errors='coerce'
        )

        validator_dict = {
            package: operator_data.loc[
                operator_data["Package Title"] == package,
                "Partner Station Id"
            ].tolist()
            for package in linear_packages
        }

        linear_packages_to_openpyxl = Workbook()
        sheet = linear_packages_to_openpyxl.active
        sheet.title = "Linear_Package"
        sheet.append(list(operator_data.columns))

        for col in range(1, 10):
            sheet.cell(row=1, column=col).fill = PatternFill(
                start_color="00CCFF", end_color="00CCFF", fill_type="solid"
            )
            sheet.column_dimensions[get_column_letter(col)].width = 25

        for row in operator_data.itertuples(index=False):
            sheet.append(list(row))

        operator_processed_files = []
        if xumo_file:
            operator_processed_files.append(xumo_file)
        if pluto_file:
            operator_processed_files.append(pluto_file)
        if plex_file:
            operator_processed_files.append(plex_file)

        for data_file in operator_processed_files:
            for linear_package_ids in linear_packages_list_ids:
                for data_file_row in data_file.iter_rows(min_row=2, values_only=True):
                    package_title = linear_package_ids["package_title"]
                    if data_file_row[1] not in validator_dict.get(package_title, []):
                        sheet.append([
                            linear_package_ids["package_id"],
                            linear_package_ids["package_title"],
                            data_file_row[6],
                            data_file_row[7],
                            data_file_row[0],
                            data_file_row[1],
                            data_file_row[2],
                            data_file_row[3],
                            data_file_row[4]
                        ])

        excel_file_path = f"{folder}/{operator_name}_linear_packages_file.xlsx"
        linear_packages_to_openpyxl.save(excel_file_path)

    else:
        print(f"Cannot generate linear packages for operator {operator_name} as it has no linear packages declared")


if __name__ == "__main__":
    pluto_partner_id = "1007223"
    plex_partner_id = "1007225"
    directories = []
    want_station_policy_map = "yes"

    # TIVO+ source files
    folder = f"TivoPlus_source_files/{sharepoint_file_name}"
    pluto_source_file,xumo_workaround_source_file,xumo_120_source_file,plex_us_source_file,xumo_ca_source_file,plex_canada_source_file = sharepoint_source_to_tivoplus_data(folder)
    pluto_ca_source_file = ""

    blacklisted_stations = create_dict_blacklisted(blacklist_location)

    for env_generator in env_generator_save_folder:
        today = str(date.today())
        directory = env_generator[1] + "_" + today
        directories.append(directory)

        if not os.path.exists(directory):
            os.makedirs(directory)
            print(directory, " Folder created")

        operators_list = generate_properties_files_in_bulk(env_generator[0])

        for operator_data in operators_list:
            operator_file = operator_data[5]
            msoServiceIDs = operator_data[6]
            msoServiceIDs_CA = operator_data[7]
            xumo_120_enabledDisabled = operator_data[8]
            station_policy = operator_data[9]
            linear_packages = operator_data[10]
            linear_packages_file = operator_data[11]
            #starting default values
            xumo_station_policy_file = 0
            pluto_file_for_packages = 0
            plex_file_for_packages = 0

            # For CableCo11, we track US/CA outputs separately to avoid overwriting
            pluto_file_for_packages_us = None
            pluto_file_for_packages_ca = None
            plex_file_for_packages_us = None
            plex_file_for_packages_ca = None
            xumo_station_policy_file_us = None
            xumo_station_policy_file_ca = None

            if "none" in operator_file.lower():
                print(f"{operator_data[0]} was skipped for Tivo+ files generation as there are no operator bulk files declared in the operators_list Excel file")

            else:
                #Pluto script execution
                if operator_data[2] != None and operator_data[2] != 101:
                    if "eastlink" in operator_data[0].lower():
                        print("Eastlink checks it")
                        if msoServiceIDs == None and operator_file == None:
                            print("Failure: " + f"{operator_data[0]} -" " cannot create operator file as both mso localities cell and bulk file are empty")
                        elif msoServiceIDs == None and operator_file != None:
                            list_msoservice_IDs = read_mso_localities("Pluto",operator_data[0], pluto_ca_packaged_service_id, operator_file, pluto_partner_id)
                        else:
                            list_msoservice_IDs = msoServiceIDs.replace("\n", ",").replace(" ", "").split(",")
                        file_processor = PlutoCA_processor(pluto_ca_source_file, pluto_partner_id, list_msoservice_IDs,operator_data[2], pluto_ca_packaged_service_id, operator_file,directory, operator_data[0])
                        file_processor.pluto_csv_to_data()
                        pluto_file_for_packages = file_processor.write_channel_page()
                        file_processor.write_msoSiteID_pages()
                        file_processor.save_file()

                    elif "cableco11" in operator_data[0].lower():
                        if (msoServiceIDs == None or msoServiceIDs_CA == None) and operator_file == None:
                            print("Failure: " + f"{operator_data[0]} -" " cannot create operator file as both mso localities cell and bulk file are empty")
                        if msoServiceIDs == None and operator_file != None:
                            list_msoservice_IDs1 = read_mso_localities("Pluto",operator_data[0], pluto_source_file, operator_file, pluto_partner_id)
                        if msoServiceIDs_CA == None and operator_file != None:
                            list_msoservice_IDs2 = read_mso_localities("Pluto",operator_data[0], pluto_ca_source_file, operator_file, pluto_partner_id)
                        else:
                            list_msoservice_IDs1 = msoServiceIDs.replace("\n", ",").replace(" ", "").split(",")
                            list_msoservice_IDs2 = msoServiceIDs_CA.replace("\n", ",").replace(" ", "").split(",")

                        if list_msoservice_IDs1 != None:
                            file_processor1 = Pluto_processor(pluto_source_file, pluto_partner_id, list_msoservice_IDs1,operator_data[2], directory, operator_data[0])
                            file_processor1.pluto_csv_to_data()
                            pluto_file_for_packages_us = file_processor1.write_channel_page()
                            file_processor1.write_msoSiteID_pages()
                            file_processor1.save_file()

                        """                        if list_msoservice_IDs2 != None:
                                                    file_processor2 = PlutoCA_processor(pluto_ca_source_file, pluto_partner_id, list_msoservice_IDs2,operator_data[2], pluto_ca_packaged_service_id, operator_file,directory, operator_data[0])
                                                    file_processor2.pluto_csv_to_data()
                                                    file_processor2.write_channel_page()
                                                    file_processor2.write_msoSiteID_pages()
                                                    file_processor2.save_file()"""

                    else:
                        if msoServiceIDs == None and operator_file == None:
                            print("Failure: " + f"{operator_data[0]} -" " cannot create operator file as both mso localities cell and bulk file are empty")
                        if msoServiceIDs == None and operator_file != None:
                            list_msoservice_IDs = read_mso_localities("Pluto", operator_data[0], pluto_packaged_service_id,operator_file,pluto_partner_id)
                            #print(list_msoservice_IDs)
                        else:
                            list_msoservice_IDs = msoServiceIDs.replace("\n", ",").replace(" ", "").split(",")
                        file_processor = Pluto_processor(pluto_source_file,pluto_partner_id,list_msoservice_IDs,operator_data[2],directory,operator_data[0])
                        file_processor.pluto_csv_to_data()
                        pluto_file_for_packages = file_processor.write_channel_page()
                        file_processor.write_msoSiteID_pages()
                        file_processor.save_file()


                #Plex script execution
                if operator_data[4] != None and operator_data[4] != 701:
                    if "eastlink" in operator_data[0].lower():
                        if msoServiceIDs == None and operator_file == None:
                            print("Failure: " + f"{operator_data[0]} -" " cannot create operator file as both mso localities cell and bulk file are empty")
                        if msoServiceIDs == None and operator_file != None:
                            list_msoservice_IDs = read_mso_localities("Plex", operator_data[0], plex_ca_packaged_service_id, operator_file, plex_partner_id,pluto="no")
                        else:
                            list_msoservice_IDs = msoServiceIDs.replace("\n", ",").replace(" ", "").split(",")
                        file_processor = PlexCA_processor(plex_canada_source_file,plex_partner_id,list_msoservice_IDs,operator_data[4],plex_ca_packaged_service_id,operator_file,directory,operator_data[0])
                        file_processor.plexca_csv_to_data()
                        file_processor.open_mso_file_data()
                        file_processor.compare_and_filter()
                        plex_file_for_packages = file_processor.write_channel_page()
                        file_processor.write_msoSiteID_pages()
                        file_processor.save_file()

                    elif "cableco11" in operator_data[0].lower():
                        if (msoServiceIDs == None or msoServiceIDs_CA == None) and operator_file == None:
                            print("Failure: " + f"{operator_data[0]} -" " cannot create operator file as both mso localities cell and bulk file are empty")
                        if msoServiceIDs == None and operator_file != None:
                            list_msoservice_IDs1 = read_mso_localities("Plex",operator_data[0], plex_us_packaged_service_id, operator_file, plex_partner_id,pluto="no")
                        if msoServiceIDs_CA == None and operator_file != None:
                            list_msoservice_IDs2 = read_mso_localities("Plex",operator_data[0], plex_ca_packaged_service_id, operator_file, plex_partner_id,pluto="no")
                        else:
                            list_msoservice_IDs1 = msoServiceIDs.replace("\n", ",").replace(" ", "").split(",")
                            list_msoservice_IDs2 = msoServiceIDs_CA.replace("\n", ",").replace(" ", "").split(",")

                        if list_msoservice_IDs1 != None:
                            file_processor1 = Plex_processor(plex_us_source_file,plex_partner_id,list_msoservice_IDs1,operator_data[4],directory,operator_data[0])
                            file_processor1.plex_csv_to_data()
                            plex_file_for_packages_us = file_processor1.write_channel_page()
                            file_processor1.write_msoSiteID_pages()
                            file_processor1.save_file()

                        if list_msoservice_IDs2 != None:
                            file_processor2 = PlexCA_processor(plex_canada_source_file,plex_partner_id,list_msoservice_IDs2,operator_data[4],plex_ca_packaged_service_id,operator_file,directory,operator_data[0])
                            file_processor2.plexca_csv_to_data()
                            file_processor2.open_mso_file_data()
                            file_processor2.compare_and_filter()
                            plex_file_for_packages_ca = file_processor2.write_channel_page()
                            file_processor2.write_msoSiteID_pages()
                            file_processor2.save_file()

                    else:
                        if msoServiceIDs == None and operator_file == None:
                            print(  "Failure: " + f"{operator_data[0]} -" " cannot create operator file as both mso localities cell and bulk file are empty")
                        if msoServiceIDs == None and operator_file != None:
                            list_msoservice_IDs = read_mso_localities("Plex",operator_data[0], plex_us_packaged_service_id, operator_file, plex_partner_id, pluto = "no")
                        else:
                            list_msoservice_IDs = msoServiceIDs.replace("\n", ",").replace(" ", "").split(",")

                        file_processor = Plex_processor(plex_us_source_file,plex_partner_id,list_msoservice_IDs,operator_data[4],directory,operator_data[0])
                        file_processor.plex_csv_to_data()
                        plex_file_for_packages = file_processor.write_channel_page()
                        file_processor.write_msoSiteID_pages()
                        file_processor.save_file()

                #Xumo script execution
                if operator_data[3] != None and operator_data[3] != 401:
                    xumo_partner_id = operator_data[1]
                    if "eastlink" in operator_data[0].lower():
                        if msoServiceIDs == None and operator_file == None:
                            print("Failure: " + f"{operator_data[0]} -" " cannot create operator file as both mso localities cell and bulk file are empty")
                        if msoServiceIDs == None and operator_file != None:
                            list_msoservice_IDs = read_mso_localities("Xumo CA", operator_data[0],xumo_ca_packaged_service_id, operator_file,xumo_partner_id)
                        else:
                            list_msoservice_IDs = msoServiceIDs.replace("\n", ",").replace(" ", "").split(",")
                        file_processor = XumoCA_processor(xumo_ca_source_file,xumo_partner_id,list_msoservice_IDs,operator_data[3],xumo_ca_packaged_service_id,operator_file,directory,operator_data[0],blacklisted_stations)
                        file_processor.xumo_csv_to_data()
                        file_processor.open_mso_file_data()
                        file_processor.compare_and_filter()
                        xumo_station_policy_file = file_processor.write_channel_page()
                        if want_station_policy_map == "yes":
                            station_policy_generator(operator_name=operator_data[0], folder=directory,station_policy_file=xumo_station_policy_file,policy_name=station_policy)
                        file_processor.write_msoSiteID_pages()
                        file_processor.save_file()

                    elif "cableco11" in operator_data[0].lower():

                        if (msoServiceIDs == None or msoServiceIDs_CA == None) and operator_file == None:
                            print("Failure: " + f"{operator_data[0]} -" " cannot create operator file as both mso localities cell and bulk file are empty")
                        if msoServiceIDs == None and operator_file != None:
                            list_msoservice_IDs1= read_mso_localities("Xumo",operator_data[0],xumo_packaged_service_id, operator_file,xumo_partner_id)
                        if msoServiceIDs_CA == None and operator_file != None:
                            list_msoservice_IDs2 = read_mso_localities("Xumo",operator_data[0],xumo_ca_packaged_service_id, operator_file,xumo_partner_id)
                        else:
                            list_msoservice_IDs1 = msoServiceIDs.replace("\n", ",").replace(" ", "").split(",")
                            list_msoservice_IDs2 = msoServiceIDs_CA.replace("\n", ",").replace(" ", "").split(",")

                        if str(xumo_120_enabledDisabled).lower() == "yes":
                            xumo_source_file = xumo_120_source_file
                        elif str(xumo_120_enabledDisabled).lower() == None:
                            xumo_source_file = xumo_workaround_source_file
                        if list_msoservice_IDs1 != None:
                            file_processor1 = Xumo_processor(xumo_source_file,xumo_partner_id,list_msoservice_IDs1,operator_data[3],xumo_packaged_service_id,operator_file,directory,operator_data[0],xumo_120_enabledDisabled)
                            file_processor1.xumo_csv_to_data()
                            file_processor1.open_mso_file_data()
                            file_processor1.compare_and_filter()
                            xumo_station_policy_file_us =file_processor1.write_channel_page()
                            if want_station_policy_map == "yes":
                                station_policy_generator(operator_name=operator_data[0], folder=directory,station_policy_file=xumo_station_policy_file_us,policy_name=station_policy)
                            file_processor1.write_msoSiteID_pages()
                            file_processor1.save_file()

                        if list_msoservice_IDs2 != None:
                            file_processor2 = XumoCA_processor(xumo_ca_source_file, xumo_partner_id,list_msoservice_IDs2, operator_data[3],xumo_ca_packaged_service_id, operator_file, directory, operator_data[0],blacklisted_stations)
                            file_processor2.xumo_csv_to_data()
                            file_processor2.open_mso_file_data()
                            file_processor2.compare_and_filter()
                            xumo_station_policy_file_ca =file_processor2.write_channel_page()
                            if want_station_policy_map == "yes":
                                station_policy_generator(operator_name=operator_data[0], folder=directory,station_policy_file=xumo_station_policy_file_ca,policy_name=station_policy)
                            file_processor2.write_msoSiteID_pages()
                            file_processor2.save_file()

                    else:
                        if msoServiceIDs == None and operator_file == None:
                            print("Failure: " + f"{operator_data[0]} -" " cannot create operator file as both mso localities cell and bulk file are empty")
                        elif msoServiceIDs == None and operator_file != None:
                            list_msoservice_IDs = read_mso_localities("Xumo",operator_data[0], xumo_packaged_service_id, operator_file,xumo_partner_id)
                        else:
                            list_msoservice_IDs = msoServiceIDs.replace("\n", ",").replace(" ", "").split(",")
                        if str(xumo_120_enabledDisabled).lower() == "yes":
                            xumo_source_file = xumo_120_source_file
                        elif str(xumo_120_enabledDisabled).lower() == "no":
                            xumo_source_file = xumo_workaround_source_file
                        #print(operator_data[0], xumo_source_file, xumo_120_enabledDisabled)
                        file_processor = Xumo_processor(xumo_source_file,xumo_partner_id,list_msoservice_IDs,operator_data[3],xumo_packaged_service_id,operator_file,directory,operator_data[0],blacklisted_stations)
                        file_processor.xumo_csv_to_data()
                        file_processor.open_mso_file_data()
                        file_processor.compare_and_filter()
                        xumo_station_policy_file = file_processor.write_channel_page()
                        if want_station_policy_map == "yes":
                            station_policy_generator(operator_name=operator_data[0],folder=directory,station_policy_file=xumo_station_policy_file,policy_name=station_policy)
                        file_processor.write_msoSiteID_pages()
                        file_processor.save_file()

                try:
                    xumo_station_policy_file
                except NameError:
                    xumo_station_policy_file = 0

                # build lists for passing to linear_packages_generator: include both US and CA workbooks when present
                pluto_files_to_pass = [f for f in (pluto_file_for_packages_us, pluto_file_for_packages_ca, pluto_file_for_packages) if f]
                plex_files_to_pass = [f for f in (plex_file_for_packages_us, plex_file_for_packages_ca, plex_file_for_packages) if f]
                xumo_files_to_pass = [f for f in (xumo_station_policy_file_us, xumo_station_policy_file_ca, xumo_station_policy_file) if f]

                # if only single file present, pass single workbook; if multiple, pass list
                pluto_arg = pluto_files_to_pass if len(pluto_files_to_pass) != 1 else pluto_files_to_pass[0] if pluto_files_to_pass else 0
                plex_arg = plex_files_to_pass if len(plex_files_to_pass) != 1 else plex_files_to_pass[0] if plex_files_to_pass else 0
                xumo_arg = xumo_files_to_pass if len(xumo_files_to_pass) != 1 else xumo_files_to_pass[0] if xumo_files_to_pass else 0

                linear_packages_generator(operator_name=operator_data[0], folder=directory, pluto_ca_packaged_service_id=pluto_ca_packaged_service_id,
                                                  plex_ca_packaged_service_id=plex_ca_packaged_service_id, xumo_ca_packaged_service_id=xumo_ca_packaged_service_id,
                                                  xumo_packaged_service_id=xumo_packaged_service_id, xumo_file=xumo_arg, pluto_file=pluto_arg, plex_file=plex_arg,
                                                  operator_linear_packages=linear_packages, operator_linear_packages_file=linear_packages_file)

    for directory in directories:
        if not os.listdir(directory):
                os.rmdir(directory)